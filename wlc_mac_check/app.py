import requests
import urllib3
import time
import re

from getpass import getpass
from netmiko import ConnectHandler
from netmiko.ssh_exception import NetMikoAuthenticationException, NetMikoTimeoutException
from openpyxl import load_workbook, Workbook

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def getVendor(mac):
    apiUrl = "https://api.macvendors.com/" + mac
    results = requests.get(apiUrl, verify=False)
    macVendor = results.content.decode('ascii')
    return macVendor

def cleanedUser(userID):
    if '\\' in userID:
        userID = userID.split("\\")[1]
    if "@" in userID:
        userID = userID.split("@")[0]
    return userID

def writeOutput(fileName):
    wb = Workbook()
    ws = wb.active
    linecount = 1
    with open("output.txt", "r") as f:
        fileData = f.readlines()
        for line in fileData:
            splitLine = re.split(r'\s{2,}', line)
            if splitLine[3] == "Unknown":
                macVendor = getVendor(splitLine[0].split()[0])
                time.sleep(1.2)
            else:
                macVendor = splitLine[3]
            try:
                ws["A{}".format(linecount)] = splitLine[0].split()[0]
                ws["B{}".format(linecount)] = splitLine[0].split()[1]
                ws["C{}".format(linecount)] = splitLine[1]
                ws["D{}".format(linecount)] = macVendor
                ws["E{}".format(linecount)] = cleanedUser(splitLine[4])
                ws["F{}".format(linecount)] = splitLine[5]
                print("{} - {}".format(linecount, splitLine[0].split()[0]))
            except IndexError:
                pass
            linecount += 1
    wb.save(fileName + '.xlsx')

def parseDeviceList(rawDeviceList):
    data = rawDeviceList.split("\n")
    with open("output.txt", "w") as f:
        for line in data:
            if ("DP" in line) or ("CORP-WIFI" in line):
                f.write(line + "\n")

def getList(userName, userPass):
    
    # on the wlc console
    # config paging disable
    # show client summary username devicetype ssid 
    # config paging enable
    
    netConnect = ConnectHandler(device_type='cisco_wlc', ip="10.11.4.15", username=userName, password=userPass)
    netConnect.send_command("config paging disable")
    deviceList = netConnect.send_command("show client summary username devicetype ssid")
    netConnect.send_command("config paging enable")
    netConnect.disconnect()
    return deviceList

userName = input("Enter your username: ")
userPass = getpass("Enter your password: ")
fileName = input("Enter file for output (xlsx will be auto aded to the end): ")

rawDeviceList = getList(userName, userPass)
parseDeviceList(rawDeviceList)
writeOutput(fileName)
